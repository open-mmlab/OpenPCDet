from openpyxl import load_workbook
import os
import numpy as np
import glob


def parse_scenes_instruction_doc(scenes_doc):
    workbook = load_workbook(scenes_doc)
    booksheet = workbook.active
    scenes_ls = []
    for i in range(7, 554):
        cell_data_1 = booksheet.cell(row=i, column=1).value
        cell_data_2 = booksheet.cell(row=i, column=2).value

        if str(cell_data_2) == '-1':
            # search all scenes which are not be annotated, 'cell_data_2' is '-1'
            print(cell_data_1, cell_data_2)
            scenes_ls.append(cell_data_1)
    return scenes_ls


def pcd2bin_32(pp, bp):
    """
    transform .pcd into .bin
    """
    rs_ls = []
    with open(pp, 'r') as f:
        raw_rs = f.readlines()[11:]
    for p_line in raw_rs:
        p_ls = p_line.strip().split(" ")
        if len(p_ls) == 4:
            if (p_ls[0] != "nan") or (p_ls[1] != "nan") or (p_ls[2] != "nan") or (p_ls[3] != "nan"):
                rs_ls.append([float(i) for i in p_line.strip().split(" ")])
            else:
                print("%s has nan:" % p_line)
        else:
            print('%s fields error!' % pp)
    np.array(rs_ls).astype(np.float32).tofile(bp)


def pcds2bins(scenes_dir, bin_p, scenes_ls):
    for scene in scenes_ls:
        pp = glob.glob(scenes_dir + scene + '/32_pcd_*')[0]
        for pcd in sorted(os.listdir(pp)):
            pcd_name = pp + '/' + pcd
            bin_name = bin_p + pcd.replace('pcd', 'bin')
            # print('pcd_name', pcd_name)
            # print('bin_name', bin_name)
            pcd2bin_32(pcd_name, bin_name)



def main():
    scenes_instuction_doc = '/nfs/neolix_data1/neolix_dataset/all_dataset/场景标注情况说明.xlsx'
    unannotated_scenes = parse_scenes_instruction_doc(scenes_instuction_doc)
    scenes_dir = '/nfs/neolix_data1/neolix_dataset/all_dataset/scenes/China/shanghai/puruan/'
    bin_dir = '/nfs/neolix_data1/neolix_dataset/develop_dataset/pseudo_label_dataset/1026_generate/bins/'
    os.makedirs(bin_dir, exist_ok=True)
    pcds2bins(scenes_dir, bin_dir, unannotated_scenes)


if __name__ == '__main__':
    main()
